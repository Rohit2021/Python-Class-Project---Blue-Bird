import socket
import sys
import time
from _thread import *
import threading 
from queue import Queue
import os
import subprocess
import openpyxl
import random

Clients = []
Names = []

print("Server intializing...\n")

def create_socket():    #creates socket
    try:
        global host
        global port
        global s
        host = "127.0.0.1"
        port = 9000
        s = socket.socket()

    except socket.error as msg:
        print("Socket creation error: " + str(msg))

def bind_socket():      #binds socket
    try:
        global host
        global port
        global s

        s.bind((host, port))
        s.listen(100)

    except socket.error as msg:
        print("Socket binding error" + str(msg) + "\n" + "Retrying...")
        bind_socket()

create_socket()
bind_socket()

def options(conn, username):
    response = ("start options function")
    conn.send(response.encode('utf-8'))
    incoming_message = conn.recv(1024) 
    incoming_message = incoming_message.decode()
    xfile = openpyxl.load_workbook('Users.xlsx')
    sheet = xfile.get_sheet_by_name('Sheet1')
    if (incoming_message == "1"):
        response = ("one")
        conn.send(response.encode('utf-8'))
    if (incoming_message == "2"):
        response = ("two")
        conn.send(response.encode('utf-8'))
        incoming_password = conn.recv(1024) 
        incoming_password = incoming_password.decode()
        file = openpyxl.load_workbook('Users.xlsx')
        sheet = xfile.get_sheet_by_name('Sheet1')
        for i in range(1000):
            username_data = sheet.cell(row = i + 1,column=2).value
            if (username == username_data):
                sheet.cell(row=i + 1, column=3, value=incoming_password)
                break
    if (incoming_message == "3"):
        response = ("three")
        conn.send(response.encode('utf-8'))
        incoming_fname = conn.recv(1024) 
        incoming_fname = incoming_fname.decode()

        incoming_lname = conn.recv(1024) 
        incoming_lname = incoming_lname.decode()

        incoming_gender = conn.recv(1024) 
        incoming_gender = incoming_gender.decode()

        incoming_interest = conn.recv(1024) 
        incoming_interest = incoming_interest.decode()

        incoming_age = conn.recv(1024) 
        incoming_age = incoming_age.decode()

        incoming_child = conn.recv(1024) 
        incoming_child = incoming_child.decode()

        incoming_ethnic = conn.recv(1024) 
        incoming_ethnic = incoming_ethnic.decode()
    
        incoming_religion = conn.recv(1024) 
        incoming_religion = incoming_religion.decode()

        incoming_school = conn.recv(1024) 
        incoming_school = incoming_school.decode()

        incoming_career = conn.recv(1024) 
        incoming_career = incoming_career.decode()

        incoming_sports = conn.recv(1024)   #changed to traveling in the excel file
        incoming_sports = incoming_sports.decode() 

        incoming_football = conn.recv(1024) 
        incoming_football = incoming_football.decode()

        incoming_baseball = conn.recv(1024) 
        incoming_baseball = incoming_baseball.decode()

        incoming_basketball = conn.recv(1024) 
        incoming_basketball = incoming_basketball.decode() 

        incoming_soccer = conn.recv(1024) 
        incoming_soccer = incoming_soccer.decode()

        incoming_other = conn.recv(1024) 
        incoming_other = incoming_other.decode()

        incoming_movies = conn.recv(1024) 
        incoming_movies = incoming_movies.decode() 

        incoming_music = conn.recv(1024) 
        incoming_music = incoming_music.decode()

        incoming_dance = conn.recv(1024) 
        incoming_dance = incoming_dance.decode()
    
        incoming_social = conn.recv(1024) 
        incoming_social = incoming_social.decode()

        incoming_reading = conn.recv(1024) 
        incoming_reading = incoming_reading.decode()

        incoming_bio = conn.recv(1024) 
        incoming_bio = incoming_bio.decode()
        for i in range(1000):
            username_data = sheet.cell(row = i + 1,column=2).value
            if (username == username_data):
                sheet.cell(row=i + 1, column=4, value=incoming_fname)
                sheet.cell(row=i + 1, column=5, value=incoming_lname)
                sheet.cell(row=i + 1, column=6, value=incoming_gender)
                sheet.cell(row=i + 1, column=7, value=incoming_interest)
                sheet.cell(row=i + 1, column=8, value=incoming_age)
                sheet.cell(row=i + 1, column=9, value=incoming_child)
                sheet.cell(row=i + 1, column=10, value=incoming_ethnic)
                sheet.cell(row=i + 1, column=11, value=incoming_religion)
                sheet.cell(row=i + 1, column=12, value=incoming_school)
                sheet.cell(row=i + 1, column=13, value=incoming_career)
                sheet.cell(row=i + 1, column=14, value=incoming_sports)
                sheet.cell(row=i + 1, column=15, value=incoming_football)
                sheet.cell(row=i + 1, column=16, value=incoming_baseball)
                sheet.cell(row=i + 1, column=17, value=incoming_basketball)
                sheet.cell(row=i + 1, column=18, value=incoming_soccer)
                sheet.cell(row=i + 1, column=19, value=incoming_other)
                sheet.cell(row=i + 1, column=20, value=incoming_movies)
                sheet.cell(row=i + 1, column=21, value=incoming_music)
                sheet.cell(row=i + 1, column=22, value=incoming_dance)
                sheet.cell(row=i + 1, column=23, value=incoming_social)
                sheet.cell(row=i + 1, column=24, value=incoming_reading)
                sheet.cell(row=i + 1, column=25, value=incoming_bio)
    if (incoming_message == "4"):
        xfile2= openpyxl.load_workbook('Issues.xlsx')
        sheet2 = xfile2.get_sheet_by_name('Sheet1')
        response = ("four")
        conn.send(response.encode('utf-8'))
        message = conn.recv(1024) 
        message = message.decode()
        for i in range(1000):
            first = sheet2.cell(row = i + 1,column=1).value
            if (first != 1):
                sheet2.cell(row=i + 1, column=1, value=1)
                sheet2.cell(row=i + 1, column=2, value=message)
                sheet2.cell(row=i + 1, column=3, value=username)
                xfile2.save('Issues.xlsx')
                break
            
    xfile.save('Users.xlsx')
        

def create_new_account(conn):
    incoming_username = conn.recv(1024) 
    incoming_username = incoming_username.decode()

    incoming_password = conn.recv(1024) 
    incoming_password = incoming_password.decode()

    incoming_fname = conn.recv(1024) 
    incoming_fname = incoming_fname.decode()

    incoming_lname = conn.recv(1024) 
    incoming_lname = incoming_lname.decode()

    incoming_gender = conn.recv(1024) 
    incoming_gender = incoming_gender.decode()

    incoming_interest = conn.recv(1024) 
    incoming_interest = incoming_interest.decode()

    incoming_age = conn.recv(1024) 
    incoming_age = incoming_age.decode()

    incoming_child = conn.recv(1024) 
    incoming_child = incoming_child.decode()

    incoming_ethnic = conn.recv(1024) 
    incoming_ethnic = incoming_ethnic.decode()

    incoming_religion = conn.recv(1024) 
    incoming_religion = incoming_religion.decode()

    incoming_school = conn.recv(1024) 
    incoming_school = incoming_school.decode()

    incoming_career = conn.recv(1024) 
    incoming_career = incoming_career.decode()

    incoming_sports = conn.recv(1024)   #changed to traveling in the excel file
    incoming_sports = incoming_sports.decode() 

    incoming_football = conn.recv(1024) 
    incoming_football = incoming_football.decode()

    incoming_baseball = conn.recv(1024) 
    incoming_baseball = incoming_baseball.decode()

    incoming_basketball = conn.recv(1024) 
    incoming_basketball = incoming_basketball.decode() 

    incoming_soccer = conn.recv(1024) 
    incoming_soccer = incoming_soccer.decode()

    incoming_other = conn.recv(1024) 
    incoming_other = incoming_other.decode()

    incoming_movies = conn.recv(1024) 
    incoming_movies = incoming_movies.decode() 

    incoming_music = conn.recv(1024) 
    incoming_music = incoming_music.decode()

    incoming_dance = conn.recv(1024) 
    incoming_dance = incoming_dance.decode()

    incoming_social = conn.recv(1024) 
    incoming_social = incoming_social.decode()

    incoming_reading = conn.recv(1024) 
    incoming_reading = incoming_reading.decode()

    incoming_bio = conn.recv(1024) 
    incoming_bio = incoming_bio.decode()

    xfile = openpyxl.load_workbook('Users.xlsx')
    sheet = xfile.get_sheet_by_name('Sheet1')
    print("\n\n\n\gupaaa")
    for i in range(1000):
        first = sheet.cell(row = i + 1,column=1).value
        if (first != 1):
            sheet.cell(row=i + 1, column=1, value=1)
            sheet.cell(row=i + 1, column=2, value=incoming_username)
            sheet.cell(row=i + 1, column=3, value=incoming_password)
            sheet.cell(row=i + 1, column=4, value=incoming_fname)
            sheet.cell(row=i + 1, column=5, value=incoming_lname)
            sheet.cell(row=i + 1, column=6, value=incoming_gender)
            sheet.cell(row=i + 1, column=7, value=incoming_interest)
            sheet.cell(row=i + 1, column=8, value=incoming_age)
            sheet.cell(row=i + 1, column=9, value=incoming_child)
            sheet.cell(row=i + 1, column=10, value=incoming_ethnic)
            sheet.cell(row=i + 1, column=11, value=incoming_religion)
            sheet.cell(row=i + 1, column=12, value=incoming_school)
            sheet.cell(row=i + 1, column=13, value=incoming_career)
            sheet.cell(row=i + 1, column=14, value=incoming_sports)
            sheet.cell(row=i + 1, column=15, value=incoming_football)
            sheet.cell(row=i + 1, column=16, value=incoming_baseball)
            sheet.cell(row=i + 1, column=17, value=incoming_basketball)
            sheet.cell(row=i + 1, column=18, value=incoming_soccer)
            sheet.cell(row=i + 1, column=19, value=incoming_other)
            sheet.cell(row=i + 1, column=20, value=incoming_movies)
            sheet.cell(row=i + 1, column=21, value=incoming_music)
            sheet.cell(row=i + 1, column=22, value=incoming_dance)
            sheet.cell(row=i + 1, column=23, value=incoming_social)
            sheet.cell(row=i + 1, column=24, value=incoming_reading)
            sheet.cell(row=i + 1, column=25, value=incoming_bio)
            break

    xfile.save('Users.xlsx')

def login(s_name, s_password):
    xfile = openpyxl.load_workbook('Users.xlsx')
    sheet = xfile.get_sheet_by_name('Sheet1')

    for i in range(1000):
        username_data = sheet.cell(row = i + 1,column=2).value
        password_data = sheet.cell(row = i + 1,column=3).value  
        if (s_password == password_data and s_name == username_data):
            return 1
            break
            
    
def send_message(conn, incoming_message, x, name):      #sends the message with the name of the sender
    message = incoming_message
    conn2 = Clients[x]
    conn2.send(message.encode('utf-8'))  #sends message

    message2 = ("From: ")   
    conn2.send(message2.encode('utf-8'))        #Adds the sender's name to the message
    message3 = (name)
    conn2.send(message3.encode('utf-8'))

    console(conn)   #returns user to console function

def get_message(conn, x):       #takes message from the user
    response = ("Enter in the message")
    conn.send(response.encode('utf-8'))
    
    incoming_message = conn.recv(1024) 
    incoming_message = incoming_message.decode()
    
    identify = ("Identify")     #Asks the client for the user's name
    conn.send(identify.encode('utf-8'))
    
    name = conn.recv(1024)      #Recieves the user's name
    name = name.decode()
    
    send_message(conn, incoming_message, x, name)
    
def message(conn):         #expands on the message command by asking the user who they wish to message
    response = ("Which user would you like to communicate with?")
    conn.send(response.encode('utf-8'))
    incoming_message = conn.recv(1024) 
    incoming_message = incoming_message.decode()

    for i in range(100):
        if incoming_message == Names[i]:
            x = i
            get_message(conn, x)
    
    response = ("Invalid Response")
    conn.send(response.encode('utf-8'))
    console(conn)   #returns user to console function

def view_profile(conn, s_name):
    xfile = openpyxl.load_workbook('Users.xlsx')
    sheet = xfile.get_sheet_by_name('Sheet1')
    for i in range(1000):
        username_data = sheet.cell(row = i + 1,column=2).value
        if (s_name == username_data):
            message = str(("start print_profile"))
            conn.send(message.encode())
            time.sleep(.00001)
            username = sheet.cell(row=i + 1, column=2).value
            username = str(username)
            conn.send(username.encode()) 
            time.sleep(.00001)
            password = sheet.cell(row=i + 1, column=3).value
            password = str(password)
            conn.send(password.encode())
            time.sleep(.00001)
            fname = sheet.cell(row=i + 1, column=4).value
            fname = str(fname)
            conn.send(fname.encode())
            time.sleep(.00001)
            lname = sheet.cell(row=i + 1, column=5).value
            lname = str(lname)
            conn.send(lname.encode())
            time.sleep(.00001)
            gender = sheet.cell(row=i + 1, column=6).value
            gender = str(gender)
            conn.send(gender.encode())
            time.sleep(.00001)
            interest = sheet.cell(row=i + 1, column=7).value
            interest = str(interest)
            conn.send(interest.encode())
            time.sleep(.00001)
            age = sheet.cell(row=i + 1, column=8).value
            age = str(age)
            conn.send(age.encode()) 
            time.sleep(.00001)
            child = sheet.cell(row=i + 1, column=9).value
            child = str(child)
            conn.send(child.encode())
            time.sleep(.00001)
            ethnic = sheet.cell(row=i + 1, column=10).value
            ethnic = str(ethnic)
            conn.send(ethnic.encode())
            time.sleep(.00001)
            religion = sheet.cell(row=i + 1, column=11).value
            religion = str(religion)
            conn.send(religion.encode())
            time.sleep(.00001)
            school = sheet.cell(row=i + 1, column=12).value
            school = str(school)
            conn.send(school.encode())
            time.sleep(.00001)
            career = sheet.cell(row=i + 1, column=13).value
            career = str(career)
            conn.send(career.encode())
            time.sleep(.00001)
            sports = sheet.cell(row=i + 1, column=14).value
            sports = str(sports)   
            conn.send(sports.encode())
            time.sleep(.00001)
            football = sheet.cell(row=i + 1, column=15).value
            football = str(football)
            conn.send(football.encode())
            time.sleep(.00001)
            baseball = sheet.cell(row=i + 1, column=16).value
            baseball = str(baseball)
            conn.send(baseball.encode())
            time.sleep(.00001)
            basketball = sheet.cell(row=i + 1, column=17).value
            basketball = str(basketball)
            conn.send(basketball.encode())
            time.sleep(.00001)
            soccer = sheet.cell(row=i + 1, column=18).value
            soccer = str(soccer)
            conn.send(soccer.encode())
            time.sleep(.00001)
            other = sheet.cell(row=i + 1, column=19).value
            other = str(other)
            conn.send(other.encode())
            time.sleep(.00001)
            movies = sheet.cell(row=i + 1, column=20).value
            movies = str(movies)
            conn.send(movies.encode())
            time.sleep(.00001)
            music = sheet.cell(row=i + 1, column=21).value
            music = str(music)
            conn.send(music.encode())
            time.sleep(.00001)
            dance = sheet.cell(row=i + 1, column=22).value
            dance = str(dance)
            conn.send(dance.encode())
            time.sleep(.00001)
            social = sheet.cell(row=i + 1, column=23).value
            social = str(social)
            conn.send(social.encode())
            time.sleep(.00001)
            reading = sheet.cell(row=i + 1, column=24).value
            reading = str(reading)
            conn.send(reading.encode())
            time.sleep(.00001)
            bio =sheet.cell(row=i + 1, column=25).value
            bio = str(bio)
            conn.send(bio.encode())
            time.sleep(.00001)

            break

def search(conn):
    message = str(("start search function"))
    conn.send(message.encode())
    username = conn.recv(1024) 
    username = username.decode() #receives username
    xfile = openpyxl.load_workbook('Users.xlsx')
    sheet = xfile.get_sheet_by_name('Sheet1')
    count = 0
    for i in range(1000):
        username_data = sheet.cell(row = i + 1,column=2).value 
        if (username == username_data):
            count = 1
            message = str(("valid"))
            conn.send(message.encode())
            time.sleep(.00001)
            view_profile(conn, username)
            break

    if (count == 0):
        message = str(("invalid"))
        conn.send(message.encode())


def matched_already(s_name, user):
    xfile = openpyxl.load_workbook('Users.xlsx')
    sheet = xfile.get_sheet_by_name('Sheet1')
    for i in range(1000):
        username = sheet.cell(row = i + 1,column=2).value
        if (s_name == username):
            for x in range(1000):
                added = sheet.cell(row = i + 1,column=x + 26).value
                if(added == user):
                    return 1
                elif(added == None):
                    return 0


def matching(conn, s_name):
    message = str(("start matching function"))
    conn.send(message.encode())
    Men = []
    Women = []
    name = ""
    xfile = openpyxl.load_workbook('Users.xlsx')
    sheet = xfile.get_sheet_by_name('Sheet1')
    interest = " "
    for i in range(1000):
        username = sheet.cell(row = i + 2,column=2).value
        interest_set = sheet.cell(row = i + 2,column=7).value
        if (s_name == username):
            interest = interest_set
            break
    for i in range(1000):
        first = sheet.cell(row = i + 2,column=1).value
        username = sheet.cell(row = i + 2,column=2).value 
        gender = sheet.cell(row = i + 2,column=6).value
        already = matched_already(s_name, username)
        if(first != 1):
            break
        if(gender == "2" and s_name != username and already == 0):
            Men.append(username)
        elif(s_name != username and already == 0):
            Women.append(username)

    if(interest == "2"):
        name = random.choice(Men)
        view_profile(conn, name)
    else:
        name = random.choice(Women)
        view_profile(conn, name)

    accept = conn.recv(1024) 
    accept = accept.decode()

    if(accept == "2"):
        for i in range(1000):
            username = sheet.cell(row = i + 1,column=2).value
            if (s_name == username):
                for x in range(100):
                    added = sheet.cell(row = i + 1,column=x + 26).value
                    if(added == None):
                        sheet.cell(row=i + 1, column=x + 26, value=name)
                        xfile.save('Users.xlsx')
                        break
                break

def message_menu(conn, s_name):
    message = str(("start messaging function"))
    conn.send(message.encode())
    xfile = openpyxl.load_workbook('Users.xlsx')
    sheet = xfile.get_sheet_by_name('Sheet1')
    info = conn.recv(1024) 
    info = info.decode()
    added_users = []
    counter = 0
    counter_inbox = 0
    if(info == "1"):
        while 1:
            for i in range(1000):
                username = sheet.cell(row = i + 1,column=2).value
                if (s_name == username):
                    for x in range(10000):
                        text = sheet.cell(row = i + 1,column=x + 250).value
                        if(text == None):
                            message = str(("stop"))
                            conn.send(message.encode())
                            break       
                        time.sleep(.00001)
                        conn.send(text.encode())
                    break
                
                
            
            receiver = conn.recv(1024) 
            receiver = receiver.decode()
            if(receiver == "exit"):         #allows user to exit the messenger area
                break
            for i in range(1000):
                username = sheet.cell(row = i + 1,column=2).value
                if (receiver == username):
                    message = str(("valid"))        #Checks to see if username entered by user is valid to receive a message
                    conn.send(message.encode())
                    counter+=1
            if (counter == 0):
                message = str(("invalid"))
                conn.send(message.encode())
                break
            else:
                message = conn.recv(1024) 
                message = message.decode()
                for i in range(1000):
                    username = sheet.cell(row = i + 1,column=2).value
                    if (receiver == username):
                        for x in range(10000):
                            inbox = sheet.cell(row = i + 1,column=x + 250).value
                            if(inbox == None):
                                sheet.cell(row=i + 1, column=x + 250, value="sent by them")
                                sheet.cell(row=i + 1, column=x + 251, value=message)       #writes the message into the mailbox on the user receiving the message
                                sheet.cell(row=i + 1, column=x + 252, value=s_name)
                                xfile.save('Users.xlsx')
                                break
                        break
                for i in range(1000):
                    username = sheet.cell(row = i + 1,column=2).value
                    if (s_name == username):
                        for x in range(10000):
                            inbox = sheet.cell(row = i + 1,column=x + 250).value
                            if(inbox == None):
                                sheet.cell(row=i + 1, column=x + 250, value="sent by you")
                                sheet.cell(row=i + 1, column=x + 251, value=message)       #writes the message into the mailbox on the user sending the message
                                sheet.cell(row=i + 1, column=x + 252, value=receiver)
                                xfile.save('Users.xlsx')
                                counter = 0
                                break
                        break
            
    elif(info == "2"):
        for i in range(1000):
            username = sheet.cell(row = i + 1,column=2).value
            if (s_name == username):
                for x in range(100):
                    added = sheet.cell(row = i + 1,column=x + 26).value
                    if(added == None):
                        break       
                    added_users.append(added)
                    counter+=1
                break
        message = str(added_users)
        conn.send(message.encode())
        time.sleep(.00001)
        message = str(counter)
        conn.send(message.encode())


 

            
    
def console(conn, s_name):      #Allows clients to enter in commands to navigate interface
    while True:
        incoming_message = conn.recv(1024) 
        incoming_message = incoming_message.decode()

        if incoming_message == "1":
            view_profile(conn, s_name)
            
        if incoming_message == "2":
            message_menu(conn, s_name)

        if incoming_message == "3":
            matching(conn, s_name)
            
        if incoming_message == "4":
            search(conn)

        if incoming_message == "5":
            options(conn, s_name)

        if incoming_message == "6":
            message = str(("start closing connection"))
            conn.send(message.encode())
            conn.close()
        
        if incoming_message == "list":
            print("\nListing connections ...\n",Names)
            response = ', '.join(Names)
            conn.send(response.encode('utf-8'))
            
        if incoming_message == "message":
            message(conn)

def close_connection():
    while True:
        data = conn.recv(1024)
        if len(data) == 0:
            conn.close()
    
def clientthread(conn):   #Takes the name of the newly entered client
    while True:
        s_account = conn.recv(1024)
        s_account = s_account.decode ()

        if (s_account == "1"):      #helps to create new account
            create_new_account(conn)

        while 1:
            s_name = conn.recv(1024)
            s_name = s_name.decode ()

            s_password = conn.recv(1024)
            s_password = s_password.decode ()

            validate_login = login(s_name, s_password)

            if (validate_login == 1):
                break
            else:
                message = str(("invalid login"))  #informs client that an incorrect login was entered
                conn.send(message.encode())
        
            
        
        message = str(("valid login"))  #informs client that a correct login was entered
        conn.send(message.encode())
        Names.append(s_name) 
        print(s_name," has joined")
        t1 = threading.Thread(target=console(conn, s_name),) 
        t2 = threading.Thread(target=close_connection(conn),)

        t1.start() 
        t2.start() 
  
        t1.join() 
        t2.join()

        #conn.send(alert.encode('utf-8'))
        #print(Names[x],"'s connection point is ", Clients[x])
        
    conn.close

def accepting_connections():    #Accepts new connections and creates new threads
    while True:
        conn, addr = s.accept()
        print("Connected with " + addr[0] + ":" +str(addr[1]))
        
        Clients.append(conn)
        start_new_thread(clientthread, (conn,))


accepting_connections()
